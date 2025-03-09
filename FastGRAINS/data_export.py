import os
import io
import base64
import zipfile
from PIL import Image
import pandas as pd
import xlsxwriter
import tempfile
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, PatternFill, numbers
import seaborn as sns
import matplotlib.pyplot as plt
from constants import OUTPUT_DIR

# Define data directories
DATA_DIR = './Results'
if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)

def export_to_excel(df, filename):
    """
    Exports a DataFrame to Excel with properly formatted cells and embedded images.
    
    Parameters:
    - df (pd.DataFrame): DataFrame containing the data to export
    - filename (str): Path where the Excel file will be saved
    
    The function handles:
    - Converting base64-encoded images to actual images in the Excel file
    - Applying proper formatting to numeric columns
    - Setting appropriate cell alignments
    """
    with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Write the header
        for col_num, column_title in enumerate(df.columns):
            cell_format = workbook.add_format({'align': 'center', 'valign': 'vcenter'})
            worksheet.write(0, col_num, column_title, cell_format)

        # Store temporary image paths to delete them later
        temp_image_paths = []

        # Write the data
        for row_num, row_data in enumerate(df.values, 1):
            for col_num, cell_value in enumerate(row_data):
                if isinstance(cell_value, str) and cell_value.startswith('<img'):
                    # Extract the base64 image data
                    base64_data = cell_value.split('base64,')[1].split('"')[0]
                    image_data = base64.b64decode(base64_data)
                    image = Image.open(io.BytesIO(image_data))
                    image = image.resize((image.width // 2, image.height // 2))  # Resize image to half
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
                        image_path = tmp.name
                        image.save(image_path)
                        temp_image_paths.append(image_path)
                    # Add the image to the cell
                    worksheet.set_row(row_num, 80)  # Adjust row height as needed
                    worksheet.set_column(col_num, col_num, 20)  # Adjust column width as needed
                    worksheet.embed_image(row_num, col_num, image_path, {'x_scale': 0.5, 'y_scale': 0.5})
                else:
                    if df.columns[col_num] in ['de', 'dins', 'dcir', 'minf', 'maxf', 'd1', 'd2']:
                        cell_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'num_format': '0'})
                    else:
                        cell_format = workbook.add_format({'align': 'center', 'valign': 'vcenter'})
                    worksheet.write(row_num, col_num, cell_value, cell_format)

        # Save the workbook
        writer.close()

        # Clean up the temporary image files
        for image_path in temp_image_paths:
            os.remove(image_path)

def shape_calculation(grains, output_excel_filename='grains_data.xlsx'):
    """
    Processes grain data to calculate shape metrics, generate images with circles,
    and export the results to an Excel file with embedded images.
    
    Parameters:
    - grains (dict): Dictionary containing grain data
    - output_excel_filename (str): Name of the output Excel file
    
    Returns:
    - new_grains_df (pd.DataFrame): DataFrame containing processed grains data
    """
    # List to collect grain IDs that encountered errors
    keys_to_delete = []

    # Iterate over grains and draw circles
    for grain_id, grain_data in grains.items():
        if not draw_circles_on_grain(grain_data):
            keys_to_delete.append(grain_id)

    # Delete grains that encountered errors
    for key in keys_to_delete:
        del grains[key]

    # Create DataFrame from grains dictionary
    grains_df = pd.DataFrame.from_dict(grains, orient='index')

    # Calculate 'dins' metric
    grains_df['dins'] = 2 * grains_df['MCI'].apply(lambda x: x[-1] if isinstance(x, (list, tuple)) else None)
    grains_df['dins'] = grains_df['dins'] * grains_df['scale_factor']

    # Adjust diameter metrics by 'scale_factor'
    for col in ['de', 'dcir', 'minf', 'maxf', 'd1', 'd2']:
        grains_df[col] = grains_df[col] / grains_df['scale_factor']

    # Select columns to include
    columns_to_include = [
        'grain_id', 'grain_results', 'grain_rs_PCD', 'de', 'dins', 'dcir',
        'minf', 'maxf', 'd1', 'd2', 'Roundness', 'AR', 'Cx',
        'Sa', 'Sc', 'Sd', 'Sp', 'Swl'
    ]
    new_grains_df = grains_df.loc[:, columns_to_include]

    # Function to convert images to HTML image tags
    def image_to_image_tag(image_obj):
        image_obj = image_obj.resize((image_obj.width // 2, image_obj.height // 2))
        with io.BytesIO() as buffer:
            image_obj.save(buffer, format="PNG")
            base64_data = base64.b64encode(buffer.getvalue()).decode('utf-8')
        return f'<img src="data:image/png;base64,{base64_data}" width="50" height="50"/>'

    # Convert images in 'grain_results' and 'grain_rs_PCD' columns
    new_grains_df['grain_results'] = new_grains_df['grain_results'].map(image_to_image_tag)
    new_grains_df['grain_rs_PCD'] = new_grains_df['grain_rs_PCD'].map(image_to_image_tag)

    # Format numerical columns to two decimal places
    columns_to_format_two_decimals = [
        'Roundness', 'Sa', 'Sd', 'Sc', 'Sp', 'Swl', 'AR', 'Cx'
    ]

    def format_two_decimals(value):
        return f'{value:.2f}'

    for col in columns_to_format_two_decimals:
        new_grains_df[col] = new_grains_df[col].map(format_two_decimals)

    # Format specific columns to zero decimal places
    columns_to_format_zero_decimals = ['de', 'dins', 'dcir', 'minf', 'maxf', 'd1', 'd2']

    def format_zero_decimals(value):
        return f'{value:.0f}'

    for col in columns_to_format_zero_decimals:
        new_grains_df[col] = new_grains_df[col].map(format_zero_decimals)

    # Convert other columns to numeric
    columns_to_convert = [
        col for col in new_grains_df.columns if col not in ['grain_id', 'grain_results', 'grain_rs_PCD']
    ]
    new_grains_df[columns_to_convert] = new_grains_df[columns_to_convert].apply(pd.to_numeric, errors='coerce')

    # Save the Excel file in DATA_DIR
    excel_path = os.path.join(DATA_DIR, output_excel_filename)
    export_to_excel(new_grains_df, excel_path)
    print(f"Results exported to {excel_path}")

    return new_grains_df

def plot_granulometric_curve(df):
    """
    Creates and saves a granulometric curve visualization from grain diameter data.
    
    Parameters:
    - df (pd.DataFrame): DataFrame containing grain data with 'd1' and 'd2' columns
    
    The function creates a log-scale plot showing the cumulative distribution of grain 
    diameters with proper styling and formatting.
    """
    # Sort the diameters and calculate cumulative percentages
    sorted_d1 = df['d1'].sort_values().reset_index(drop=True)
    sorted_d2 = df['d2'].sort_values().reset_index(drop=True)
    cumulative_percentages = (sorted_d1.rank(method='first') / len(sorted_d1)) * 100

    sns.set(style="whitegrid")

    # Use a different color palette
    colors = sns.color_palette("hsv", 2)

    # Plot the granulometric curve
    plt.figure(figsize=(5, 5))
    plt.plot(sorted_d1, cumulative_percentages, label='d1 (px)', color=colors[0])
    plt.plot(sorted_d2, cumulative_percentages, label='d2 (px)', color=colors[1])
    plt.fill_between(sorted_d2, cumulative_percentages, cumulative_percentages, color=colors[1], alpha=0.2)
    plt.xlabel('Diameter (px)', fontweight='bold', color='black')
    plt.ylabel('Cumulative Percentage (%)', fontweight='bold', color='black')
    plt.title('Granulometric Curve', fontweight='bold', color='black')
    plt.legend()
    plt.tight_layout()

    plt.grid(False)
    plt.minorticks_on()
    plt.tick_params(axis='both', which='both', direction='in', color='black')
    plt.tick_params(axis='both', which='major', length=6, width=2)
    plt.tick_params(axis='both', which='minor', length=4, width=1)

    plt.gca().spines['bottom'].set_color('black')
    plt.gca().spines['left'].set_color('black')
    plt.gca().spines['top'].set_color('none')
    plt.gca().spines['right'].set_color('none')

    # Add graduation marks
    plt.gca().xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x:.3f}'))
    plt.gca().yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f'{y:.0f}'))

    plt.xscale('log')
    plt.gca().invert_xaxis()

    # Save the plot
    granulometric_curve_path = os.path.join(DATA_DIR, 'granulometric_curve.png')
    plt.savefig(granulometric_curve_path, dpi=300)
    plt.show()
    plt.close()

def plot_density_curves(df):
    """
    Creates and saves density curve visualizations for shape metrics.
    
    Parameters:
    - df (pd.DataFrame): DataFrame containing grain shape metrics
    
    The function creates kernel density estimation plots for key shape metrics
    (Roundness, AR, Cx, Sp) with proper styling, annotations showing mean and
    standard deviation, and saves the result to the output directory.
    """
    variables = ['Roundness', 'AR', 'Cx', 'Sp']
    labels = ['$R$', '$AR$', '$C_{x}$', '$S_{p}$']
    letters = ['(a)', '(b)', '(c)', '(d)']

    # Use a Seaborn color palette
    colors = sns.color_palette("husl", len(variables))

    # Configure font and axes properties
    plt.rcParams.update({
        'font.size': 9.5,
        'axes.labelsize': 8.5,
        'xtick.labelsize': 8.5,
        'ytick.labelsize': 8.5,
        'legend.fontsize': 5.5,
        'font.family': 'DejaVu Sans',
        'axes.edgecolor': 'black',
        'xtick.color': 'black',
        'ytick.color': 'black'
    })

    # Create subplots
    fig, axes = plt.subplots(2, 2, figsize=(5, 4), dpi=220)
    axes = axes.flatten()

    # Plot density curves with shaded areas
    for i, (variable, label, color) in enumerate(zip(variables, labels, colors)):
        sns.kdeplot(df[variable], fill=True, color=color, ax=axes[i], label=variable, alpha=0.3, linestyle='-')

        axes[i].set_xlabel(label)
        axes[i].set_ylabel('Density')
        axes[i].set_xlim(0, 1)  # Set x-axis range between 0 and 1
        axes[i].grid(False)  # Turn off the grid

        # Calculate mean and standard deviation
        mean = df[variable].mean()
        std = df[variable].std()

        # Create text with mean and standard deviation
        textstr = f'{variable} = {mean:.2f}, \n σ = {std:.2f}'

        # Add text in a fancy box
        props = dict(boxstyle='round,pad=0.3', edgecolor='black', facecolor='white', alpha=0.95)
        axes[i].text(0.05, 0.95, textstr, transform=axes[i].transAxes, fontsize=8,
                    verticalalignment='top', bbox=props)

        # Position the letter outside the plot to the left
        axes[i].text(-0.2, 1.05, f'{letters[i]}', transform=axes[i].transAxes, fontsize=10, fontweight='bold', va='top', ha='right')

    # Adjust layout
    plt.tight_layout()

    # Save the plot
    density_curves_path = os.path.join(OUTPUT_DIR, 'density_curves.png')
    plt.savefig(density_curves_path, dpi=300)
    plt.show()
    plt.close()

def save_data_as_zip(material_name):
    """
    Compresses all files in the OUTPUT_DIR into a zip file.
    
    Parameters:
    - material_name (str): Name used as prefix for the zip file
    
    The function creates a zip archive containing all files from the OUTPUT_DIR,
    preserving the relative path structure.
    """
    zip_filename = material_name + '.zip'
    with zipfile.ZipFile(zip_filename, 'w') as zipf:
        for root, dirs, files_in in os.walk(OUTPUT_DIR):
            for file in files_in:
                file_path = os.path.join(root, file)
                zipf.write(file_path, arcname=os.path.relpath(file_path, OUTPUT_DIR))
    print(f"Data saved to {zip_filename}")

def export_to_excel(material_name, grains):
    """
    Creates a formatted Excel report for grain analysis with embedded images.
    
    Parameters:
    - material_name (str): Name used for the output file
    - grains (dict): Dictionary containing grain data
    
    The function:
    1. Converts the grains dictionary to a DataFrame
    2. Generates density curve plots
    3. Creates an Excel file with proper formatting
    4. Embeds images into the Excel cells
    5. Adds units and styling to improve readability
    """
    # Convert the grains dictionary to a DataFrame
    grains_df = pd.DataFrame.from_dict(grains, orient='index')

    # Select columns to include
    columns_to_include = [
        'grain_id', 'grain_rs', 'R_image', 'de', 'dcir',
        'minf', 'maxf', 'Roundness', 'AR', 'Cx', 'Sp', "SAGI"
    ]

    new_grains_df = grains_df.loc[:, columns_to_include]

    # Plot density curves
    plot_density_curves(new_grains_df)

    filename = os.path.join(OUTPUT_DIR, material_name + ".xlsx")

    # Save the DataFrame to Excel
    new_grains_df.to_excel(filename, index=False)

    # Load the workbook and worksheet
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # Add a second header row for units
    units_row = [
        '',   # grain_id (no units)
        '',   # grain_rs (no units)
        '',   # R_image (no units)
        'px', # de (pixels)
        'px', # dcir (pixels)
        'px', # minf (pixels)
        'px', # maxf (pixels)
        '-',  # Roundness (dimensionless)
        '-',  # AR (dimensionless)
        '-',  # Cx (dimensionless)
        '-',  # Sp (dimensionless)
        '-'   # SAGI (dimensionless)
    ]

    # Insert the units row below the main header
    ws.insert_rows(2)  # Insert a new row at position 2
    for col_num, unit in enumerate(units_row, 1):
        ws.cell(row=2, column=col_num, value=unit)

    # Function to insert images from a specific column
    def insert_images_from_column(column_name, new_width=120):
        # Find the column index for the header
        img_col_index = None
        for cell in ws[1]:  # Iterate over the header row
            if cell.value == column_name:
                img_col_index = cell.column  # Get the column index (1-based)
                break

        if img_col_index is None:
            print(f"Column '{column_name}' not found.")
        else:
            # Iterate over the cells in the column (starting from row 3, since row 2 is the units row)
            for row in ws.iter_rows(min_row=3, min_col=img_col_index, max_col=img_col_index):
                cell = row[0]
                # Get the corresponding PIL Image from the DataFrame
                img_pil = new_grains_df.loc[cell.row - 3, column_name]  # DataFrame is 0-based, Excel is 1-based

                # Convert PIL Image to bytes
                img_bytes = io.BytesIO()
                img_pil.save(img_bytes, format='PNG')
                img_bytes.seek(0)

                # Create an openpyxl Image object
                img_obj = OpenpyxlImage(img_bytes)

                # Optionally, adjust image size
                original_width, original_height = img_obj.width, img_obj.height
                scale = new_width / original_width
                new_height = int(original_height * scale)
                img_obj.width = new_width
                img_obj.height = new_height

                # Adjust the row height: Excel row height is measured in points
                # Roughly, 1 point ≈ 1.33 pixels, so set height = new_height / 1.33
                ws.row_dimensions[cell.row].height = new_height / 1.33

                # Adjust the column width to fit the image
                ws.column_dimensions[openpyxl.utils.get_column_letter(img_col_index)].width = new_width / 7  # Approximate conversion

                # Clear the cell text so only the image is visible
                cell.value = None

                # Anchor the image to the cell
                img_obj.anchor = cell.coordinate
                ws.add_image(img_obj)

    # Insert images from 'R_image' column
    insert_images_from_column('R_image', new_width=120)

    # Insert images from 'grain_rs' column
    insert_images_from_column('grain_rs', new_width=120)

    # Apply formatting to all cells
    center_alignment = Alignment(horizontal='center', vertical='center')
    decimal_format = numbers.FORMAT_NUMBER_00  # Format for two decimal places

    # Style for the header
    header_font = Font(bold=True, color="FFFFFF")  # White text
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Black background

    # Style for the units row
    units_font = Font(italic=True, color="FFFFFF")  # White italic text
    units_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Black background

    # Apply formatting to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_alignment  # Center align all cells
            if isinstance(cell.value, (int, float)):  # Format numeric cells to two decimal places
                cell.number_format = decimal_format

    # Apply header styling
    for cell in ws[1]:  # Iterate over the main header row
        cell.font = header_font
        cell.fill = header_fill

    # Apply units row styling
    for cell in ws[2]:  # Iterate over the units row
        cell.font = units_font
        cell.fill = units_fill

    # Save the modified workbook
    wb.save(filename)

    print(f"Excel file '{filename}' has been created with images anchored to cells, formatting, and units applied.")

# Note: draw_circles_on_grain function is referenced but not defined in this file
# You will need to import it from the appropriate module
from visualization import draw_circles_on_grain
